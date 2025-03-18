#!/usr/bin/env python3
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.utils import get_column_letter
import re
import html

class AdvancedExcelConverter:
    """
    Advanced Excel to Markdown converter that handles complex Excel features:
    - Merged cells
    - Cell formatting (bold, italic, etc.)
    - Cell alignments
    - Basic formula display
    - Cell colors and backgrounds (as notes in the Markdown)
    """
    
    def __init__(self, excel_file):
        """Initialize with Excel file path"""
        self.excel_file = excel_file
        self.workbook = openpyxl.load_workbook(excel_file, data_only=False)
        self.workbook_data_only = openpyxl.load_workbook(excel_file, data_only=True)
    
    def get_sheet_names(self):
        """Get all sheet names from the workbook"""
        return self.workbook.sheetnames
    
    def _is_cell_bold(self, cell):
        """Check if a cell has bold formatting"""
        return cell.font.bold
    
    def _is_cell_italic(self, cell):
        """Check if a cell has italic formatting"""
        return cell.font.italic
    
    def _get_cell_alignment(self, cell):
        """Get cell alignment (left, center, right)"""
        if cell.alignment.horizontal == 'center':
            return 'center'
        elif cell.alignment.horizontal == 'right':
            return 'right'
        else:
            return 'left'  # Default
    
    def _get_merged_cell_ranges(self, sheet):
        """Get all merged cell ranges in the sheet"""
        return sheet.merged_cells.ranges
    
    def _is_cell_in_merged_range(self, cell_coord, merged_ranges):
        """Check if a cell is part of a merged range"""
        for merged_range in merged_ranges:
            if cell_coord in merged_range:
                return merged_range
        return None
    
    def _format_value(self, value, is_bold=False, is_italic=False):
        """Format cell value with Markdown styling"""
        if value is None:
            return ""
            
        # Convert to string and escape any Markdown special characters
        value_str = str(value)
        value_str = value_str.replace('|', '\\|')
        value_str = value_str.replace('\n', '<br>')
        
        # Apply formatting
        if is_bold:
            value_str = f"**{value_str}**"
        if is_italic:
            value_str = f"*{value_str}*"
            
        return value_str
    
    def convert_sheet_to_markdown(self, sheet_name, include_formulas=False):
        """
        Convert a specific sheet to Markdown with advanced formatting
        
        Parameters:
        -----------
        sheet_name : str
            Name of the sheet to convert
        include_formulas : bool, optional
            Whether to include formulas as comments in the output
            
        Returns:
        --------
        str
            Markdown representation of the sheet
        """
        sheet = self.workbook[sheet_name]
        sheet_data = self.workbook_data_only[sheet_name]
        merged_ranges = self._get_merged_cell_ranges(sheet)
        
        # Find the actual data range (skip completely empty rows/columns)
        data_rows = []
        max_cols = 0
        
        for row_idx, row in enumerate(sheet.iter_rows(), 1):
            row_has_data = False
            for col_idx, cell in enumerate(row, 1):
                if cell.value is not None:
                    row_has_data = True
                    max_cols = max(max_cols, col_idx)
            if row_has_data:
                data_rows.append(row_idx)
        
        if not data_rows:
            return f"## {sheet_name}\n\n*Empty sheet*\n\n"
        
        min_row, max_row = min(data_rows), max(data_rows)
        
        # Prepare the Markdown table
        column_alignments = []
        for col_idx in range(1, max_cols + 1):
            cell = sheet.cell(min_row, col_idx)
            align = self._get_cell_alignment(cell)
            column_alignments.append(align)
        
        # Create header row
        md_table = "| "
        for col_idx in range(1, max_cols + 1):
            cell = sheet.cell(min_row, col_idx)
            cell_data = sheet_data.cell(min_row, col_idx)
            
            # Check if this cell is part of a merged range
            cell_coord = f"{get_column_letter(col_idx)}{min_row}"
            merged_range = self._is_cell_in_merged_range(cell_coord, merged_ranges)
            
            value = cell_data.value
            is_bold = self._is_cell_bold(cell)
            is_italic = self._is_cell_italic(cell)
            
            if merged_range and cell_coord != merged_range.coord:
                # This is a continuation of a merged cell, leave it empty
                md_table += " | "
            else:
                cell_text = self._format_value(value, is_bold, is_italic)
                md_table += f"{cell_text} | "
        
        md_table += "\n|"
        
        # Create alignment row
        for align in column_alignments:
            if align == 'center':
                md_table += " :---: |"
            elif align == 'right':
                md_table += " ---: |"
            else:
                md_table += " :--- |"
        
        md_table += "\n"
        
        # Create data rows
        for row_idx in range(min_row + 1, max_row + 1):
            if row_idx in data_rows:
                md_table += "| "
                for col_idx in range(1, max_cols + 1):
                    cell = sheet.cell(row_idx, col_idx)
                    cell_data = sheet_data.cell(row_idx, col_idx)
                    
                    # Check if this cell is part of a merged range
                    cell_coord = f"{get_column_letter(col_idx)}{row_idx}"
                    merged_range = self._is_cell_in_merged_range(cell_coord, merged_ranges)
                    
                    value = cell_data.value
                    formula = cell.value if isinstance(cell.value, str) and cell.value.startswith('=') else None
                    is_bold = self._is_cell_bold(cell)
                    is_italic = self._is_cell_italic(cell)
                    
                    if merged_range and cell_coord != merged_range.coord:
                        # This is a continuation of a merged cell, leave it empty
                        md_table += " | "
                    else:
                        cell_text = self._format_value(value, is_bold, is_italic)
                        
                        # Add formula as comment if requested
                        if include_formulas and formula:
                            cell_text += f" <!-- Formula: {html.escape(formula)} -->"
                            
                        md_table += f"{cell_text} | "
                
                md_table += "\n"
        
        return md_table
    
    def convert_to_markdown(self, sheet_name=None, include_formulas=False):
        """
        Convert Excel file to Markdown
        
        Parameters:
        -----------
        sheet_name : str or list, optional
            Sheet name(s) to convert. If None, all sheets are converted.
        include_formulas : bool, optional
            Whether to include formulas as comments in the output
            
        Returns:
        --------
        str
            Markdown representation of the Excel file
        """
        # Get sheet names to process
        if sheet_name is None:
            sheet_names = self.get_sheet_names()
        elif isinstance(sheet_name, str):
            sheet_names = [sheet_name]
        else:
            sheet_names = sheet_name
        
        # Convert each sheet
        md_content = ""
        for sheet in sheet_names:
            if len(sheet_names) > 1:
                md_content += f"## {sheet}\n\n"
            
            md_content += self.convert_sheet_to_markdown(sheet, include_formulas)
            md_content += "\n\n"
        
        return md_content


def convert_excel_advanced(excel_file, output_file=None, sheet_name=None, include_formulas=False):
    """
    Convert Excel file to Markdown with advanced formatting
    
    Parameters:
    -----------
    excel_file : str
        Path to the Excel file
    output_file : str, optional
        Path to the output Markdown file. If None, return as string.
    sheet_name : str or list, optional
        Sheet name(s) to convert. If None, all sheets are converted.
    include_formulas : bool, optional
        Whether to include formulas as comments in the output
        
    Returns:
    --------
    str or bool
        If output_file is None, returns Markdown content as string.
        Otherwise, returns True if successful, False if error occurred.
    """
    try:
        converter = AdvancedExcelConverter(excel_file)
        md_content = converter.convert_to_markdown(sheet_name, include_formulas)
        
        if output_file:
            with open(output_file, 'w', encoding='utf-8') as f:
                f.write(md_content)
            return True
        else:
            return md_content
            
    except Exception as e:
        print(f"Error in advanced conversion: {e}")
        return False


if __name__ == "__main__":
    import sys
    
    if len(sys.argv) < 2:
        print("Usage: python advanced_converter.py <excel_file> [output_file] [sheet_name]")
        sys.exit(1)
    
    excel_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else None
    sheet_name = sys.argv[3] if len(sys.argv) > 3 else None
    
    result = convert_excel_advanced(excel_file, output_file, sheet_name)
    
    if output_file:
        if result:
            print(f"Converted Excel file to Markdown: {output_file}")
        else:
            print("Conversion failed.")
    else:
        print(result) 